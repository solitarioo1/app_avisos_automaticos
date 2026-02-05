"""
Rutas de Avisos - API endpoints para gestión de avisos meteorológicos
"""
import csv
import json
import logging
import os
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path

from flask import (Blueprint, Response, jsonify, render_template, request,
                   stream_with_context)

BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / 'OUTPUT'
logger = logging.getLogger(__name__)
active_processes = {}

avisos_bp = Blueprint('avisos', __name__, url_prefix='')


@avisos_bp.route('/avisos', methods=['GET'])
def avisos():
    """Página de gestión de avisos - Conectado a BD o archivos locales"""
    avisos_lista = []
    aviso_param = request.args.get('aviso')
    filtro_numero = int(aviso_param) if aviso_param else None

    try:
        import psycopg2
        import psycopg2.extras

        conn = psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            port=int(os.getenv("DB_PORT", "5432")),
            database=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD")
        )
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        query = (
            "SELECT DISTINCT numero_aviso, fecha_emision, titulo, nivel, "
            "color FROM avisos_completos WHERE color IN ('rojo', 'naranja') "
            "ORDER BY numero_aviso DESC"
        )
        cursor.execute(query)
        avisos_bd = cursor.fetchall()
        cursor.close()
        conn.close()

        for aviso in avisos_bd:
            numero = aviso['numero_aviso']
            if filtro_numero and numero != filtro_numero:
                continue

            json_path = BASE_DIR / 'JSON' / 'aviso_{}.json'.format(numero)
            output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)

            estado_descargado = json_path.exists()
            mapas_creados = (output_path.exists() and
                           (any(output_path.glob('*.webp')) or
                            any(output_path.glob('*.png'))))

            css_class = ('table-success' if mapas_creados else
                        ('table-warning' if estado_descargado else ''))
            avisos_lista.append({
                'numero': numero,
                'titulo': aviso['titulo'],
                'nivel': aviso['nivel'],
                'color': aviso['color'],
                'fecha_emision': str(aviso['fecha_emision']),
                'descargado': '✅' if estado_descargado else '⏳',
                'mapa_creado': '✅' if mapas_creados else '⏳',
                'estado_css': css_class
            })
    except (psycopg2.Error, ImportError):
        logger.warning("BD no disponible, usando JSON locales")
        json_dir = BASE_DIR / 'JSON'
        if json_dir.exists():
            for json_file in sorted(json_dir.glob('aviso_*.json'),
                                   reverse=True):
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        numero_str = data.get('numero_aviso',
                                            json_file.stem.replace('aviso_',
                                                                  ''))
                        numero = int(numero_str)

                        if filtro_numero and numero != filtro_numero:
                            continue

                        color = data.get('color', 'plomo').lower()
                        if color not in ['rojo', 'naranja']:
                            continue

                        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)
                        mapas_creados = (output_path.exists() and
                                       (any(output_path.glob('*.webp')) or
                                        any(output_path.glob('*.png'))))

                        avisos_lista.append({
                            'numero': numero,
                            'titulo': data.get('titulo', 'Aviso {}'.format(
                                numero)),
                            'nivel': data.get('nivel', 'AMARILLO').upper(),
                            'color': color,
                            'fecha_emision': data.get('fecha_emision',
                                                     '2026-02-01'),
                            'descargado': '✅',
                            'mapa_creado': '✅' if mapas_creados else '⏳',
                            'estado_css': ('table-success' if mapas_creados
                                         else 'table-warning')
                        })
                except (ValueError, KeyError, json.JSONDecodeError,
                       OSError):
                    pass

    return render_template('avisos.html', avisos=avisos_lista)


@avisos_bp.route('/api/avisos/<int:numero>/descargar', methods=['POST'])
def api_descargar_aviso(numero):
    """API para descargar JSON de aviso desde SENAMHI"""
    try:
        json_path = BASE_DIR / 'JSON' / 'aviso_{}.json'.format(numero)

        if json_path.exists():
            return jsonify({
                'success': True,
                'message': 'Aviso ya descargado',
                'file': str(json_path)
            }), 200

        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'descargar_aviso.py'),
             str(numero)],
            capture_output=True,
            text=True,
            timeout=60,
            check=False
        )

        if result.returncode == 0:
            if json_path.exists():
                return jsonify({
                    'success': True,
                    'message': 'Aviso {} descargado correctamente'.format(
                        numero),
                    'file': str(json_path)
                }), 200
            return jsonify({
                'success': False,
                'error': 'Descarga completada pero archivo no encontrado'
            }), 400
        return jsonify({
            'success': False,
            'error': 'Error en descarga: {}'.format(result.stderr)
        }), 400

    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'Timeout: La descarga tardó demasiado'
        }), 408
    except OSError as e:
        logger.error("Error descargando aviso %d: %s", numero, str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos/<int:numero>/procesar',
                  methods=['POST', 'GET'])
def api_procesar_aviso(numero):
    """API para generar mapas del aviso con soporte Server-Sent Events"""
    try:
        stream = request.args.get('stream', 'false').lower() == 'true'
        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)

        if output_path.exists() and (any(output_path.glob('*.webp')) or
                                     any(output_path.glob('*.png'))):
            if stream:
                def generate_existing():
                    msg = {'type': 'log', 'message': 'Mapas ya existen',
                           'severity': 'info'}
                    yield "data: {}\n\n".format(json.dumps(msg))
                    msg2 = {'type': 'complete',
                            'message': 'Procesamiento completado'}
                    yield "data: {}\n\n".format(json.dumps(msg2))
                return Response(stream_with_context(generate_existing()),
                              mimetype='text/event-stream')
            return jsonify({
                'success': True,
                'message': 'Mapas ya existen',
                'path': str(output_path)
            }), 200

        if stream:
            def generate():
                process = None
                try:
                    msg_init = {
                        'type': 'log',
                        'message': ('Iniciando procesamiento del aviso '
                                   '{}...').format(numero),
                        'severity': 'info'
                    }
                    yield "data: {}\n\n".format(json.dumps(msg_init))

                    process = subprocess.Popen(
                        [sys.executable,
                         str(BASE_DIR / 'procesar_aviso.py'),
                         str(numero)],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        bufsize=1,
                        universal_newlines=True
                    )

                    active_processes[numero] = process

                    for line in process.stdout:
                        if (numero in active_processes and
                            active_processes[numero] is None):
                            process.terminate()
                            break

                        line = line.rstrip('\n')
                        if line:
                            if 'ERROR' in line or 'Error' in line:
                                sev = 'error'
                            elif ('SUCCESS' in line or 'completado' in line
                                  or 'generado' in line):
                                sev = 'success'
                            else:
                                sev = 'info'

                            msg = {'type': 'log', 'message': line,
                                   'severity': sev}
                            yield "data: {}\n\n".format(json.dumps(msg))

                    for line in process.stderr:
                        line = line.rstrip('\n')
                        if line:
                            msg = {'type': 'log',
                                   'message': '[STDERR] {}'.format(line),
                                   'severity': 'warning'}
                            yield "data: {}\n\n".format(json.dumps(msg))

                    returncode = process.wait()

                    if returncode == 0:
                        img_files = (list(output_path.glob('*.webp')) +
                                    list(output_path.glob('*.png')))
                        num_mapas = len(img_files)
                        msg_succ = {
                            'type': 'log',
                            'message': ('✅ {} mapas generados '
                                       'exitosamente').format(num_mapas),
                            'severity': 'success'
                        }
                        yield "data: {}\n\n".format(json.dumps(msg_succ))
                        msg_comp = {
                            'type': 'complete',
                            'message': ('Procesamiento completado - {} '
                                       'mapas').format(num_mapas)
                        }
                        yield "data: {}\n\n".format(json.dumps(msg_comp))
                    else:
                        msg_err = {
                            'type': 'error',
                            'message': ('Proceso terminó con código '
                                       '{}').format(returncode)
                        }
                        yield "data: {}\n\n".format(json.dumps(msg_err))
                        msg_fail = {'type': 'complete',
                                   'message': 'Procesamiento fallido'}
                        yield "data: {}\n\n".format(json.dumps(msg_fail))

                    if numero in active_processes:
                        del active_processes[numero]

                except OSError as e:
                    msg_ex = {'type': 'error', 'message': str(e)}
                    yield "data: {}\n\n".format(json.dumps(msg_ex))
                    msg_ex2 = {'type': 'complete', 'message': 'Error'}
                    yield "data: {}\n\n".format(json.dumps(msg_ex2))
                    if numero in active_processes:
                        del active_processes[numero]

            response = Response(stream_with_context(generate()),
                              mimetype='text/event-stream')
            response.headers['Cache-Control'] = 'no-cache'
            response.headers['X-Accel-Buffering'] = 'no'
            return response

        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'procesar_aviso.py'),
             str(numero)],
            capture_output=True,
            text=True,
            timeout=300,
            check=False
        )

        if result.returncode == 0:
            img_files = (list(output_path.glob('*.webp')) +
                        list(output_path.glob('*.png')))
            csv_files = list(output_path.glob('*.csv'))

            return jsonify({
                'success': True,
                'message': 'Mapas generados correctamente',
                'mapas': len(img_files),
                'csvs': len(csv_files),
                'path': str(output_path)
            }), 200

        return jsonify({
            'success': False,
            'error': 'Error en procesamiento: {}'.format(result.stderr)
        }), 400

    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'Timeout: El procesamiento tardó demasiado'
        }), 408
    except OSError as e:
        logger.error("Error procesando aviso %d: %s", numero, str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos/<int:numero>/cancel', methods=['POST'])
def api_cancel_aviso(numero):
    """API para cancelar la generación de mapas en curso"""
    try:
        if numero in active_processes:
            process = active_processes[numero]
            if process and process.poll() is None:
                import signal
                if hasattr(signal, 'CTRL_C_EVENT'):
                    process.send_signal(signal.CTRL_C_EVENT)
                else:
                    process.terminate()
                    try:
                        process.wait(timeout=5)
                    except subprocess.TimeoutExpired:
                        process.kill()

                active_processes[numero] = None
                logger.info("Proceso de aviso %d cancelado", numero)
                return jsonify({
                    'success': True,
                    'message': 'Generación del aviso {} cancelada'.format(
                        numero)
                }), 200

        return jsonify({
            'success': False,
            'error': 'Proceso no encontrado o ya finalizado'
        }), 404

    except OSError as e:
        logger.error("Error cancelando aviso %d: %s", numero, str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos/<int:numero>/exportar-excel',
                  methods=['POST'])
def api_exportar_excel(numero):
    """API para generar Excel con datos de departamentos/provincias"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)
        excel_path = output_path / 'aviso_{}_reporte.xlsx'.format(numero)

        if excel_path.exists():
            return jsonify({
                'success': True,
                'message': 'Excel ya existe',
                'file': str(excel_path)
            }), 200

        wb = Workbook()

        if (output_path / 'distritos_afectados.csv').exists():
            ws_distritos = wb.active
            ws_distritos.title = 'Distritos'

            with open(output_path / 'distritos_afectados.csv', 'r',
                     encoding='utf-8') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_distritos.cell(row=row_idx,
                                               column=col_idx,
                                               value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            fill = PatternFill(start_color="0070C0",
                                             end_color="0070C0",
                                             fill_type="solid")
                            cell.fill = fill
                        cell.alignment = Alignment(horizontal="left",
                                                  vertical="center")

        if (output_path / 'provincias_afectadas.csv').exists():
            ws_provincias = wb.create_sheet('Provincias')

            with open(output_path / 'provincias_afectadas.csv', 'r',
                     encoding='utf-8') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_provincias.cell(row=row_idx,
                                                column=col_idx,
                                                value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            fill = PatternFill(start_color="00B050",
                                             end_color="00B050",
                                             fill_type="solid")
                            cell.fill = fill
                        cell.alignment = Alignment(horizontal="left",
                                                  vertical="center")

        for ws_name in wb.sheetnames:
            ws_obj = wb[ws_name]
            for column in ws_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                    except (ValueError, TypeError):
                        pass
                ws_obj.column_dimensions[column_letter].width = (
                    min(max_length + 2, 50))

        wb.save(str(excel_path))

        return jsonify({
            'success': True,
            'message': 'Excel generado correctamente',
            'file': str(excel_path)
        }), 200

    except OSError as e:
        logger.error("Error generando Excel para aviso %d: %s", numero,
                    str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/mapas/aviso/<int:numero>', methods=['GET'])
def api_mapas_por_aviso(numero):
    """API para obtener lista de mapas de un aviso"""
    try:
        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)
        mapas = []

        if output_path.exists():
            for img_file in output_path.glob('*.*'):
                if img_file.suffix.lower() in ['.webp', '.png']:
                    mapas.append({
                        'nombre': img_file.stem,
                        'archivo': img_file.name,
                        'url': '/mapas/imagen/{}/{}'.format(
                            numero, img_file.name),
                        'ruta': str(img_file)
                    })

        return jsonify({
            'success': True,
            'aviso': numero,
            'mapas': mapas,
            'cantidad': len(mapas)
        }), 200

    except OSError as e:
        logger.error("Error obteniendo mapas de aviso %d: %s", numero,
                    str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos/<int:numero>/info', methods=['GET'])
def api_info_aviso(numero):
    """API para obtener info del aviso (nivel, departamentos afectados)"""
    try:
        import psycopg2
        import psycopg2.extras

        json_path = BASE_DIR / 'JSON' / 'aviso_{}.json'.format(numero)
        datos = None
        color = 'plomo'

        if json_path.exists():
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                    color = json_data.get('color', 'plomo')
                    datos = json_data
            except (ValueError, json.JSONDecodeError, OSError):
                pass

        if not datos:
            try:
                conn = psycopg2.connect(
                    host=os.getenv("DB_HOST", "localhost"),
                    port=int(os.getenv("DB_PORT", "5432")),
                    database=os.getenv("DB_NAME"),
                    user=os.getenv("DB_USER"),
                    password=os.getenv("DB_PASSWORD")
                )
                cursor = conn.cursor(
                    cursor_factory=psycopg2.extras.RealDictCursor)

                query = (
                    "SELECT numero_aviso, titulo, nivel FROM "
                    "avisos_completos WHERE numero_aviso = %s LIMIT 1"
                )
                cursor.execute(query, (numero,))
                result = cursor.fetchone()
                cursor.close()
                conn.close()

                if result:
                    datos = {
                        'numero_aviso': result['numero_aviso'],
                        'titulo': result['titulo'],
                        'nivel': result['nivel']
                    }
            except (psycopg2.Error, ImportError):
                pass

        if not datos:
            return jsonify({
                'success': False,
                'error': 'Aviso no encontrado'
            }), 404

        deptos = set()
        for dia in range(1, 4):
            key = 'dep_afectados_dia{}'.format(dia)
            if key in datos and datos[key]:
                deptos.update([d.strip() for d in datos[key].split(',')])

        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)
        mapas_creados = (output_path.exists() and
                        (any(output_path.glob('*.webp')) or
                         any(output_path.glob('*.png'))))

        return jsonify({
            'success': True,
            'numero': numero,
            'titulo': datos.get('titulo', ''),
            'nivel': datos.get('nivel', ''),
            'color': color,
            'departamentos': list(deptos) if deptos else [],
            'mapas_creados': mapas_creados,
            'fecha_emision': datos.get('fecha_emision', '')
        }), 200

    except OSError as e:
        logger.error("Error obteniendo info de aviso %d: %s", numero, str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos/<int:numero>/departamentos',
                  methods=['GET'])
def api_departamentos_aviso(numero):
    """API para obtener departamentos afectados del CSV"""
    try:
        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)
        csv_path = output_path / 'distritos_afectados.csv'

        mapas_creados = (output_path.exists() and
                        (any(output_path.glob('*.webp')) or
                         any(output_path.glob('*.png'))))

        if not mapas_creados:
            return jsonify({
                'success': True,
                'departamentos': [],
                'mapas_creados': False,
                'message': 'Mapas no creados'
            }), 200

        departamentos = set()
        if csv_path.exists():
            try:
                with open(csv_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        depto = row.get('DEPARTAMEN', '').strip()
                        if depto:
                            departamentos.add(depto)
            except (OSError, ValueError):
                pass

        return jsonify({
            'success': True,
            'departamentos': sorted(list(departamentos)),
            'mapas_creados': mapas_creados,
            'csv_path': str(csv_path)
        }), 200

    except OSError as e:
        logger.error("Error obteniendo departamentos para aviso %d: %s",
                    numero, str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos', methods=['GET'])
def api_avisos():
    """API para obtener lista de avisos desde BD y OUTPUT/"""
    try:
        import psycopg2
        import psycopg2.extras

        avisos_dict = {}

        try:
            conn = psycopg2.connect(
                host=os.getenv("DB_HOST", "localhost"),
                port=int(os.getenv("DB_PORT", "5432")),
                database=os.getenv("DB_NAME"),
                user=os.getenv("DB_USER"),
                password=os.getenv("DB_PASSWORD")
            )
            cursor = conn.cursor(
                cursor_factory=psycopg2.extras.RealDictCursor)

            query = (
                "SELECT DISTINCT numero_aviso, titulo, nivel, color, "
                "fecha_emision FROM avisos_completos WHERE color IN "
                "('rojo', 'naranja') ORDER BY numero_aviso DESC"
            )
            cursor.execute(query)
            avisos_bd = cursor.fetchall()
            cursor.close()
            conn.close()

            for aviso in avisos_bd:
                numero = aviso['numero_aviso']
                avisos_dict[numero] = {
                    'numero': numero,
                    'titulo': aviso['titulo'],
                    'nivel': aviso['nivel'],
                    'color': aviso.get('color', 'plomo'),
                    'fecha_emision': str(aviso.get('fecha_emision', '')),
                    'descargado': '✅',
                    'mapa_creado': '⏳',
                    'fuente': 'bd'
                }
        except (psycopg2.Error, ImportError):
            pass

        if (BASE_DIR / 'JSON').exists():
            for json_file in sorted((BASE_DIR / 'JSON').glob('aviso_*.json'),
                                   reverse=True):
                try:
                    numero_str = json_file.stem.split('_')[1]
                    numero = int(numero_str)

                    if numero not in avisos_dict:
                        with open(json_file, 'r', encoding='utf-8') as f:
                            datos = json.load(f)

                        color = datos.get('color', 'plomo')
                        if color.lower() not in ['rojo', 'naranja']:
                            continue

                        output_path = OUTPUT_DIR / 'aviso_{}'.format(numero)
                        mapas_creados = (output_path.exists() and
                                       (any(output_path.glob('*.webp')) or
                                        any(output_path.glob('*.png'))))

                        avisos_dict[numero] = {
                            'numero': numero,
                            'titulo': datos.get('titulo',
                                              'Aviso {}'.format(numero)),
                            'nivel': datos.get('nivel', 'AMARILLO'),
                            'color': color,
                            'fecha_emision': datos.get('fecha_emision',
                                                      '2026-02-01'),
                            'descargado': '✅',
                            'mapa_creado': '✅' if mapas_creados else '⏳',
                            'fuente': 'json'
                        }
                except (ValueError, KeyError, json.JSONDecodeError,
                       OSError):
                    pass

        if OUTPUT_DIR.exists():
            for carpeta in OUTPUT_DIR.iterdir():
                if carpeta.is_dir() and carpeta.name.startswith('aviso_'):
                    try:
                        numero_str = carpeta.name.split('_')[1]
                        numero = int(numero_str)
                        if numero not in avisos_dict:
                            has_maps = (any(carpeta.glob('*.webp')) or
                                       any(carpeta.glob('*.png')))
                            avisos_dict[numero] = {
                                'numero': numero,
                                'titulo': 'Aviso {}'.format(numero),
                                'nivel': 'N/A',
                                'color': 'plomo',
                                'fecha_emision': '2026-02-01',
                                'descargado': '⏳',
                                'mapa_creado': '✅' if has_maps else '⏳',
                                'fuente': 'output'
                            }
                    except (ValueError, OSError):
                        pass

        avisos_sorted = sorted(avisos_dict.values(),
                       key=lambda x: x['numero'],
                       reverse=True)

        return jsonify({
            'success': True,
            'avisos': avisos_sorted
        }), 200

    except OSError as e:
        logger.error("Error obteniendo avisos: %s", str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@avisos_bp.route('/api/avisos/nuevos', methods=['GET'])
def api_avisos_nuevos():
    """API para obtener avisos nuevos en las últimas 24 horas"""
    try:
        import psycopg2
        import psycopg2.extras

        conn = psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            port=int(os.getenv("DB_PORT", "5432")),
            database=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            connect_timeout=5
        )
        cursor = conn.cursor(
            cursor_factory=psycopg2.extras.RealDictCursor)

        hace_24_horas = datetime.now() - timedelta(hours=24)

        query = (
            "SELECT COUNT(DISTINCT numero_aviso) as total "
            "FROM avisos_completos WHERE color IN ('rojo', 'naranja') "
            "AND fecha_emision >= %s"
        )
        cursor.execute(query, (hace_24_horas,))
        resultado = cursor.fetchone()
        total_nuevos = resultado['total'] if resultado else 0

        cursor.close()
        conn.close()

        return jsonify({
            'status': 'success',
            'total_nuevos': total_nuevos,
            'fecha_consulta': datetime.now().isoformat()
        })
    except psycopg2.OperationalError:
        logger.error("Error de conexión BD")
        return jsonify({
            'status': 'error',
            'message': ('No hay conexión a la base de datos. Intenta más '
                       'tarde.')
        }), 503
    except (psycopg2.Error, ImportError):
        logger.error("Error al obtener avisos nuevos")
        return jsonify({
            'status': 'error',
            'message': 'Error al consultar avisos nuevos'
        }), 500
