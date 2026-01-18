"""
API Flask para procesamiento automático de avisos SENAMHI
Interfaz HTTP para integración con n8n + Dashboard Web
"""

from flask import Flask, request, jsonify, send_from_directory, render_template, Response, stream_with_context
from pathlib import Path
import sys
import os
import logging
from datetime import datetime
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Agregar LAYOUT al path para importar procesar_aviso
sys.path.insert(0, str(Path(__file__).parent))

# Importaciones condicionales para evitar errores de dependencias
try:
    from procesar_aviso import procesar_aviso
    PROCESAR_AVISO_DISPONIBLE = True
except ImportError as e:
    logger.warning(f"Módulo procesar_aviso no disponible: {e}")
    PROCESAR_AVISO_DISPONIBLE = False
    def procesar_aviso(*args, **kwargs):
        return {"success": False, "error": "Módulo no disponible"}

try:
    from CONFIG.db import obtener_aviso_por_numero
    DB_DISPONIBLE = True
except ImportError as e:
    logger.warning(f"Módulo CONFIG.db no disponible: {e}")
    DB_DISPONIBLE = False
    def obtener_aviso_por_numero(numero):
        return None

# Inicializar Flask
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

# Excluir directorios del watchdog de desarrollo para evitar reinicios
app.config['DONT_RELOAD_REGEX'] = r'(\.git|__pycache__|\.pytest_cache|node_modules|TEMP|OUTPUT|\.egg-info)'

# Rutas base (desde .env)
BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / os.getenv('OUTPUT_DIR', 'OUTPUT')
DOMAIN = os.getenv('DOMAIN', 'https://mapas.miagentepersonal.me')

# Diccionario global para rastrear procesos en ejecución
active_processes = {}

# ============================================================================
# RUTAS WEB - PÁGINAS PRINCIPALES
# ============================================================================

@app.route('/', methods=['GET'])
def dashboard():
    """Dashboard principal"""
    try:
        # Obtener estadísticas
        stats = obtener_estadisticas()
        
        # Obtener avisos recientes
        avisos_recientes = obtener_avisos_recientes()
        
        # Obtener mapas recientes
        mapas_recientes = obtener_mapas_recientes()
        
        # Obtener evento actual
        evento_actual = obtener_evento_actual()
        
        return render_template('dashboard.html', 
                             stats=stats,
                             avisos_recientes=avisos_recientes,
                             mapas_recientes=mapas_recientes,
                             evento_actual=evento_actual)
    except Exception as e:
        logger.error(f"Error en dashboard: {str(e)}")
        return render_template('dashboard.html', 
                             stats={}, avisos_recientes=[], mapas_recientes=[])

@app.route('/avisos', methods=['GET'])
def avisos():
    """Página de gestión de avisos - Conectado a BD"""
    try:
        import psycopg2
        import psycopg2.extras
        
        # Obtener avisos de BD
        conn = psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            port=int(os.getenv("DB_PORT", "5432")),
            database=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD")
        )
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        query = """
        SELECT DISTINCT numero_aviso, fecha_emision, titulo, nivel, color
        FROM avisos_completos
        WHERE color IN ('rojo', 'naranja')
        ORDER BY numero_aviso DESC
        """
        cursor.execute(query)
        avisos_bd = cursor.fetchall()
        cursor.close()
        conn.close()
        
        # Procesar avisos y verificar estado
        avisos_lista = []
        for aviso in avisos_bd:
            numero = aviso['numero_aviso']
            json_path = BASE_DIR / 'JSON' / f'aviso_{numero}.json'
            output_path = OUTPUT_DIR / f'aviso_{numero}'
            
            estado_descargado = json_path.exists()
            # Verificar si existen mapas (webp o png)
            mapas_creados = output_path.exists() and (any(output_path.glob('*.webp')) or any(output_path.glob('*.png')))
            
            avisos_lista.append({
                'numero': numero,
                'titulo': aviso['titulo'],
                'nivel': aviso['nivel'],
                'color': aviso['color'],
                'fecha_emision': str(aviso['fecha_emision']),
                'descargado': '✅' if estado_descargado else '⏳',
                'mapa_creado': '✅' if mapas_creados else '⏳',
                'estado_css': 'table-success' if mapas_creados else ('table-warning' if estado_descargado else '')
            })
        
        return render_template('avisos.html', avisos=avisos_lista)
    except Exception as e:
        logger.error(f"Error en página avisos: {str(e)}")
        return render_template('avisos.html', avisos=[])

@app.route('/api/avisos/<int:numero>/descargar', methods=['POST'])
def api_descargar_aviso(numero):
    """API para descargar JSON de aviso desde SENAMHI"""
    try:
        import subprocess
        
        json_path = BASE_DIR / 'JSON' / f'aviso_{numero}.json'
        
        # Si ya existe, no descargar de nuevo
        if json_path.exists():
            return jsonify({
                'success': True,
                'message': 'Aviso ya descargado',
                'file': str(json_path)
            }), 200
        
        # Ejecutar descargar_aviso.py
        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'descargar_aviso.py'), str(numero)],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode == 0:
            if json_path.exists():
                return jsonify({
                    'success': True,
                    'message': f'Aviso {numero} descargado correctamente',
                    'file': str(json_path)
                }), 200
            else:
                return jsonify({
                    'success': False,
                    'error': 'Descarga completada pero archivo no encontrado'
                }), 400
        else:
            return jsonify({
                'success': False,
                'error': f'Error en descarga: {result.stderr}'
            }), 400
            
    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'Timeout: La descarga tardó demasiado'
        }), 408
    except Exception as e:
        logger.error(f"Error descargando aviso {numero}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/avisos/<int:numero>/procesar', methods=['POST', 'GET'])
def api_procesar_aviso(numero):
    """API para generar mapas del aviso con soporte para Server-Sent Events"""
    try:
        import subprocess
        import json
        
        # Verificar si se solicita stream (SSE)
        stream = request.args.get('stream', 'false').lower() == 'true'
        
        output_path = OUTPUT_DIR / f'aviso_{numero}'
        
        # Si ya existe, mapas ya están generados
        if output_path.exists() and (any(output_path.glob('*.webp')) or any(output_path.glob('*.png'))):
            if stream:
                def generate():
                    yield f"data: {json.dumps({'type': 'log', 'message': 'Mapas ya existen', 'severity': 'info'})}\n\n"
                    yield f"data: {json.dumps({'type': 'complete', 'message': 'Procesamiento completado'})}\n\n"
                return Response(stream_with_context(generate()), mimetype='text/event-stream')
            else:
                return jsonify({
                    'success': True,
                    'message': 'Mapas ya existen',
                    'path': str(output_path)
                }), 200
        
        if stream:
            # Modo stream - SSE
            def generate():
                process = None
                try:
                    import time
                    import threading
                    
                    yield f"data: {json.dumps({'type': 'log', 'message': f'Iniciando procesamiento del aviso {numero}...', 'severity': 'info'})}\n\n"
                    
                    # Ejecutar procesar_aviso.py con captura de salida
                    process = subprocess.Popen(
                        [sys.executable, str(BASE_DIR / 'procesar_aviso.py'), str(numero)],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        bufsize=1,
                        universal_newlines=True
                    )
                    
                    # Guardar proceso en diccionario global
                    active_processes[numero] = process
                    
                    # Leer stdout línea por línea
                    for line in process.stdout:
                        # Verificar si se solicitó cancelación
                        if numero in active_processes and active_processes[numero] is None:
                            process.terminate()
                            break
                        
                        line = line.rstrip('\n')
                        if line:
                            # Parsear logs del procesar_aviso.py
                            if 'ERROR' in line or 'Error' in line:
                                yield f"data: {json.dumps({'type': 'log', 'message': line, 'severity': 'error'})}\n\n"
                            elif 'SUCCESS' in line or 'completado' in line or 'generado' in line:
                                yield f"data: {json.dumps({'type': 'log', 'message': line, 'severity': 'success'})}\n\n"
                            elif 'Processing' in line or 'Procesando' in line:
                                # Intentar extraer progreso
                                yield f"data: {json.dumps({'type': 'log', 'message': line, 'severity': 'info'})}\n\n"
                            else:
                                yield f"data: {json.dumps({'type': 'log', 'message': line, 'severity': 'info'})}\n\n"
                    
                    # Leer stderr si hay
                    for line in process.stderr:
                        line = line.rstrip('\n')
                        if line:
                            yield f"data: {json.dumps({'type': 'log', 'message': f'[STDERR] {line}', 'severity': 'warning'})}\n\n"
                    
                    # Esperar a que termine el proceso
                    returncode = process.wait()
                    
                    if returncode == 0:
                        # Contar archivos generados
                        img_files = list(output_path.glob('*.webp')) + list(output_path.glob('*.png'))
                        
                        yield f"data: {json.dumps({'type': 'log', 'message': f'✅ {len(img_files)} mapas generados exitosamente', 'severity': 'success'})}\n\n"
                        yield f"data: {json.dumps({'type': 'complete', 'message': f'Procesamiento completado - {len(img_files)} mapas'})}\n\n"
                    else:
                        yield f"data: {json.dumps({'type': 'error', 'message': f'Proceso terminó con código {returncode}'})}\n\n"
                        yield f"data: {json.dumps({'type': 'complete', 'message': 'Procesamiento fallido'})}\n\n"
                    
                    # Limpiar diccionario
                    if numero in active_processes:
                        del active_processes[numero]
                        
                except Exception as e:
                    logger.error(f"Error en SSE para aviso {numero}: {str(e)}")
                    yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
                    yield f"data: {json.dumps({'type': 'complete', 'message': 'Error'})}\n\n"
                    if numero in active_processes:
                        del active_processes[numero]
            
            response = Response(stream_with_context(generate()), mimetype='text/event-stream')
            response.headers['Cache-Control'] = 'no-cache'
            response.headers['X-Accel-Buffering'] = 'no'
            return response
        
        else:
            # Modo tradicional - POST
            result = subprocess.run(
                [sys.executable, str(BASE_DIR / 'procesar_aviso.py'), str(numero)],
                capture_output=True,
                text=True,
                timeout=300
            )
            
            if result.returncode == 0:
                img_files = list(output_path.glob('*.webp')) + list(output_path.glob('*.png'))
                csv_files = list(output_path.glob('*.csv'))
                
                return jsonify({
                    'success': True,
                    'message': f'Mapas generados correctamente',
                    'mapas': len(img_files),
                    'csvs': len(csv_files),
                    'path': str(output_path)
                }), 200
            else:
                return jsonify({
                    'success': False,
                    'error': f'Error en procesamiento: {result.stderr}'
                }), 400
            
    except subprocess.TimeoutExpired:
        if stream:
            def generate_error():
                yield f"data: {json.dumps({'type': 'error', 'message': 'Timeout: El procesamiento tardó demasiado'})}\n\n"
                yield f"data: {json.dumps({'type': 'complete', 'message': 'Timeout'})}\n\n"
            return Response(stream_with_context(generate_error()), mimetype='text/event-stream')
        else:
            return jsonify({
                'success': False,
                'error': 'Timeout: El procesamiento tardó demasiado'
            }), 408
    except Exception as e:
        logger.error(f"Error procesando aviso {numero}: {str(e)}")
        if stream:
            def generate_error():
                yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
                yield f"data: {json.dumps({'type': 'complete', 'message': 'Error'})}\n\n"
            return Response(stream_with_context(generate_error()), mimetype='text/event-stream')
        else:
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500

@app.route('/api/avisos/<int:numero>/cancel', methods=['POST'])
def api_cancel_aviso(numero):
    """API para cancelar la generación de mapas en curso"""
    try:
        if numero in active_processes:
            process = active_processes[numero]
            if process and process.poll() is None:  # Proceso aún en ejecución
                # Terminar el proceso
                import signal
                if hasattr(signal, 'CTRL_C_EVENT'):  # Windows
                    process.send_signal(signal.CTRL_C_EVENT)
                else:  # Unix
                    process.terminate()
                    process.wait(timeout=5)
                
                # Marcar como cancelado
                active_processes[numero] = None
                
                logger.info(f"Proceso de aviso {numero} cancelado")
                return jsonify({
                    'success': True,
                    'message': f'Generación del aviso {numero} cancelada'
                }), 200
        
        return jsonify({
            'success': False,
            'error': 'Proceso no encontrado o ya finalizado'
        }), 404
        
    except Exception as e:
        logger.error(f"Error cancelando aviso {numero}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/avisos/<int:numero>/exportar-excel', methods=['POST'])
def api_exportar_excel(numero):
    """API para generar Excel con datos de departamentos/provincias afectadas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import csv
        
        output_path = OUTPUT_DIR / f'aviso_{numero}'
        excel_path = output_path / f'aviso_{numero}_reporte.xlsx'
        
        # Si ya existe, retornar ruta
        if excel_path.exists():
            return jsonify({
                'success': True,
                'message': 'Excel ya existe',
                'file': str(excel_path)
            }), 200
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja de Distritos
        if (output_path / 'distritos_afectados.csv').exists():
            ws_distritos = wb.active
            ws_distritos.title = 'Distritos'
            
            with open(output_path / 'distritos_afectados.csv', 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_distritos.cell(row=row_idx, column=col_idx, value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Hoja de Provincias
        if (output_path / 'provincias_afectadas.csv').exists():
            ws_provincias = wb.create_sheet('Provincias')
            
            with open(output_path / 'provincias_afectadas.csv', 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_provincias.cell(row=row_idx, column=col_idx, value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar ancho de columnas
        for ws_name in wb.sheetnames:
            ws_obj = wb[ws_name]
            for column in ws_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_obj.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(str(excel_path))
        
        return jsonify({
            'success': True,
            'message': 'Excel generado correctamente',
            'file': str(excel_path)
        }), 200
        
    except Exception as e:
        logger.error(f"Error generando Excel para aviso {numero}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/mapas/aviso/<int:numero>', methods=['GET'])
def api_mapas_por_aviso(numero):
    """API para obtener lista de mapas de un aviso"""
    try:
        output_path = OUTPUT_DIR / f'aviso_{numero}'
        mapas = []
        
        if output_path.exists():
            # Buscar mapas (webp y png)
            for img_file in output_path.glob('*.*'):
                if img_file.suffix.lower() in ['.webp', '.png']:
                    # Extraer nombre del departamento
                    depto = img_file.stem
                    mapas.append({
                        'nombre': depto,
                        'archivo': img_file.name,
                        'url': f'/mapas/imagen/{numero}/{img_file.name}',
                        'ruta': str(img_file)
                    })
        
        return jsonify({
            'success': True,
            'aviso': numero,
            'mapas': mapas,
            'cantidad': len(mapas)
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo mapas de aviso {numero}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/avisos/<int:numero>/info', methods=['GET'])
def api_info_aviso(numero):
    """API para obtener info del aviso (nivel, departamentos afectados)"""
    try:
        import json
        import psycopg2
        import psycopg2.extras
        
        json_path = BASE_DIR / 'JSON' / f'aviso_{numero}.json'
        datos = None
        color = 'plomo'  # default
        
        # 1. SIEMPRE obtener color del JSON si existe
        if json_path.exists():
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                    color = json_data.get('color', 'plomo')
                    datos = json_data
            except Exception as e:
                logger.warning(f"Error leyendo JSON {numero}: {e}")
        
        # 2. Si no está en JSON, obtener de BD
        if not datos:
            try:
                conn = psycopg2.connect(
                    host=os.getenv("DB_HOST", "localhost"),
                    port=int(os.getenv("DB_PORT", "5432")),
                    database=os.getenv("DB_NAME"),
                    user=os.getenv("DB_USER"),
                    password=os.getenv("DB_PASSWORD")
                )
                cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                
                query = "SELECT numero_aviso, titulo, nivel FROM avisos_completos WHERE numero_aviso = %s LIMIT 1"
                cursor.execute(query, (numero,))
                result = cursor.fetchone()
                cursor.close()
                conn.close()
                
                if result:
                    datos = {
                        'numero_aviso': result['numero_aviso'],
                        'titulo': result['titulo'],
                        'nivel': result['nivel']
                    }
            except Exception as e:
                logger.warning(f"No se encontró en BD para aviso {numero}: {e}")
        
        if not datos:
            return jsonify({
                'success': False,
                'error': 'Aviso no encontrado'
            }), 404
        
        # Obtener departamentos afectados (concatenar días)
        deptos = set()
        for dia in range(1, 4):
            key = f'dep_afectados_dia{dia}'
            if key in datos and datos[key]:  # Verificar que no sea None
                deptos.update([d.strip() for d in datos[key].split(',')])
        
        output_path = OUTPUT_DIR / f'aviso_{numero}'
        mapas_creados = output_path.exists() and (any(output_path.glob('*.webp')) or any(output_path.glob('*.png')))
        
        return jsonify({
            'success': True,
            'numero': numero,
            'titulo': datos.get('titulo', ''),
            'nivel': datos.get('nivel', ''),
            'color': color,  # Siempre del JSON
            'departamentos': list(deptos) if deptos else [],
            'mapas_creados': mapas_creados,
            'fecha_emision': datos.get('fecha_emision', '')
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo info de aviso {numero}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/avisos/<int:numero>/departamentos', methods=['GET'])
def api_departamentos_aviso(numero):
    """API para obtener departamentos afectados del CSV (solo si mapas están creados)"""
    try:
        import csv
        
        output_path = OUTPUT_DIR / f'aviso_{numero}'
        csv_path = output_path / 'distritos_afectados.csv'
        
        # Verificar si los mapas están creados
        mapas_creados = output_path.exists() and (any(output_path.glob('*.webp')) or any(output_path.glob('*.png')))
        
        if not mapas_creados:
            # Si no hay mapas creados, NO retornar departamentos
            return jsonify({
                'success': True,
                'departamentos': [],
                'mapas_creados': False,
                'message': 'Mapas no creados'
            }), 200
        
        # Leer CSV si existe
        departamentos = set()
        if csv_path.exists():
            try:
                with open(csv_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        depto = row.get('DEPARTAMEN', '').strip()
                        if depto:
                            departamentos.add(depto)
            except Exception as e:
                logger.warning(f"Error leyendo CSV {numero}: {e}")
        
        return jsonify({
            'success': True,
            'departamentos': sorted(list(departamentos)),
            'mapas_creados': mapas_creados,
            'csv_path': str(csv_path)
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo departamentos para aviso {numero}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/avisos', methods=['GET'])
def api_avisos():
    """API para obtener lista de avisos desde BD y OUTPUT/"""
    try:
        import psycopg2
        import psycopg2.extras
        import json
        
        avisos_dict = {}
        
        # 1. Obtener desde BD
        try:
            conn = psycopg2.connect(
                host=os.getenv("DB_HOST", "localhost"),
                port=int(os.getenv("DB_PORT", "5432")),
                database=os.getenv("DB_NAME"),
                user=os.getenv("DB_USER"),
                password=os.getenv("DB_PASSWORD")
            )
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            query = "SELECT DISTINCT numero_aviso, titulo, nivel, color FROM avisos_completos WHERE color IN ('rojo', 'naranja') ORDER BY numero_aviso DESC"
            cursor.execute(query)
            avisos_bd = cursor.fetchall()
            cursor.close()
            conn.close()
            
            for aviso in avisos_bd:
                numero = aviso['numero_aviso']
                avisos_dict[numero] = {
                    'numero': numero,
                    'titulo': aviso['titulo'],
                    'nivel': aviso['nivel'],
                    'fuente': 'bd'
                }
        except Exception as e:
            logger.warning(f"No se pudo conectar a BD: {e}")
        
        # 2. Buscar carpetas en OUTPUT/ que no estén en BD
        if OUTPUT_DIR.exists():
            for carpeta in OUTPUT_DIR.iterdir():
                if carpeta.is_dir() and carpeta.name.startswith('aviso_'):
                    numero = int(carpeta.name.split('_')[1])
                    
                    if numero not in avisos_dict:
                        # Intentar obtener info del JSON
                        json_path = BASE_DIR / 'JSON' / f'aviso_{numero}.json'
                        if json_path.exists():
                            try:
                                with open(json_path, 'r', encoding='utf-8') as f:
                                    datos = json.load(f)
                                avisos_dict[numero] = {
                                    'numero': numero,
                                    'titulo': datos.get('titulo', f'Aviso {numero}'),
                                    'nivel': datos.get('nivel', 'N/A'),
                                    'fuente': 'json'
                                }
                            except:
                                avisos_dict[numero] = {
                                    'numero': numero,
                                    'titulo': f'Aviso {numero}',
                                    'nivel': 'N/A',
                                    'fuente': 'output'
                                }
                        else:
                            avisos_dict[numero] = {
                                'numero': numero,
                                'titulo': f'Aviso {numero}',
                                'nivel': 'N/A',
                                'fuente': 'output'
                            }
        
        avisos = sorted(avisos_dict.values(), key=lambda x: x['numero'], reverse=True)
        
        return jsonify({
            'success': True,
            'avisos': avisos
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo avisos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/mapas/imagen/<int:numero>/<filename>')
def servir_mapa(numero, filename):
    """Sirve imágenes de mapas"""
    try:
        directorio = OUTPUT_DIR / f'aviso_{numero}'
        return send_from_directory(directorio, filename)
    except Exception as e:
        logger.error(f"Error sirviendo imagen: {str(e)}")
        return "Archivo no encontrado", 404

@app.route('/mapas', methods=['GET'])
def mapas():
    """Galería de mapas"""
    try:
        aviso_filtro = request.args.get('aviso')
        mapas_lista = obtener_lista_mapas(aviso_filtro)
        return render_template('mapas.html', mapas=mapas_lista)
    except Exception as e:
        logger.error(f"Error en página mapas: {str(e)}")
        return render_template('mapas.html', mapas=[])

@app.route('/decisiones')
def decisiones():
    """Página de toma de decisiones - Centro de comando"""
    try:
        stats = obtener_estadisticas()
        evento_actual = obtener_evento_actual()
        
        return render_template('decisiones.html',
                             stats=stats,
                             evento_actual=evento_actual)
    except Exception as e:
        logger.error(f"Error en página decisiones: {str(e)}")
        return render_template('decisiones.html', stats={}, evento_actual=None)

@app.route('/whatsapp', methods=['GET'])
def whatsapp():
    """Página de WhatsApp masivo"""
    try:
        stats_wa = obtener_stats_whatsapp()
        avisos_disponibles = obtener_avisos_para_whatsapp()
        historial = obtener_historial_whatsapp()
        contactos = obtener_contactos_recientes()
        
        return render_template('whatsapp.html', 
                             stats=stats_wa,
                             avisos_disponibles=avisos_disponibles,
                             historial_envios=historial,
                             contactos_recientes=contactos)
    except Exception as e:
        logger.error(f"Error en página whatsapp: {str(e)}")
        return render_template('whatsapp.html', 
                             stats={}, avisos_disponibles=[], 
                             historial_envios=[], contactos_recientes=[])

@app.route('/configuracion', methods=['GET'])
def configuracion():
    """Página de configuración"""
    return render_template('configuracion.html')

@app.route('/logs', methods=['GET'])
def logs():
    """Página de logs del sistema"""
    return render_template('logs.html')

def obtener_estadisticas():
    """Obtener estadísticas generales del sistema"""
    return {
        'avisos_activos': 3,
        'departamentos_afectados': 7,
        'agricultores_afectados': 1247,
        'agricultores_registrados': 15420,
        'mensajes_enviados': 2847,
        'cultivos_riesgo': 12,
        'probabilidad_evento': 85
    }

def obtener_evento_actual():
    """Obtener información del evento meteorológico actual"""
    return {
        'titulo': 'Lluvias Intensas y Vientos Fuertes',
        'tipo': 'Precipitaciones Pluviales',
        'vigencia': '22 Dic 2024 - 25 Dic 2024',
        'nivel_alerta': 'AMARILLO',
        'probabilidad': 85,
        'departamentos_afectados': ['Lima', 'Arequipa', 'Cusco', 'Piura', 'La Libertad', 'Lambayeque', 'Ica'],
        'cultivos_riesgo': ['Arroz', 'Maíz', 'Papa', 'Quinua', 'Algodón', 'Café', 'Cacao', 'Frijol', 'Trigo', 'Cebada', 'Palto', 'Mango'],
        'intensidad': 'Moderada a Fuerte',
        'descripcion': 'Se esperan precipitaciones de moderada a fuerte intensidad que podrían afectar las actividades agrícolas en las zonas indicadas.'
    }

def obtener_avisos_recientes():
    """Obtener lista de avisos recientes"""
    avisos = []
    
    if OUTPUT_DIR.exists():
        avisos_dirs = sorted([d for d in OUTPUT_DIR.iterdir() 
                            if d.is_dir() and d.name.startswith('aviso_')], 
                           key=lambda x: x.stat().st_mtime, reverse=True)[:10]
        
        for aviso_dir in avisos_dirs:
            numero = aviso_dir.name.replace('aviso_', '')
            mapas = len([f for f in aviso_dir.iterdir() if f.suffix == '.webp'])
            
            avisos.append({
                'numero': numero,
                'fecha': datetime.fromtimestamp(aviso_dir.stat().st_mtime).strftime('%d/%m/%Y %H:%M'),
                'estado_texto': 'Procesado',
                'estado_color': 'success',
                'num_mapas': mapas
            })
    
    return avisos

def obtener_mapas_recientes():
    """Obtener mapas recientes para preview"""
    mapas = []
    
    if OUTPUT_DIR.exists():
        avisos_dirs = sorted([d for d in OUTPUT_DIR.iterdir() 
                            if d.is_dir() and d.name.startswith('aviso_')], 
                           key=lambda x: x.stat().st_mtime, reverse=True)[:5]
        
        for aviso_dir in avisos_dirs:
            numero_aviso = aviso_dir.name.replace('aviso_', '')
            webp_files = [f for f in aviso_dir.iterdir() if f.suffix == '.webp']
            
            for webp_file in webp_files[:2]:  # Máximo 2 mapas por aviso
                mapas.append({
                    'nombre': webp_file.stem,
                    'numero_aviso': numero_aviso,
                    'url': f"{DOMAIN}/OUTPUT/{aviso_dir.name}/{webp_file.name}",
                    'thumbnail': f"{DOMAIN}/OUTPUT/{aviso_dir.name}/{webp_file.name}"
                })
    
    return mapas[:8]  # Máximo 8 mapas en total

# ============================================================================
# FUNCIONES AUXILIARES PARA PÁGINAS
# ============================================================================

def obtener_lista_avisos_completa():
    """Obtener lista completa de avisos para página de avisos"""
    avisos = []
    
    if OUTPUT_DIR.exists():
        avisos_dirs = sorted([d for d in OUTPUT_DIR.iterdir() 
                            if d.is_dir() and d.name.startswith('aviso_')], 
                           key=lambda x: x.stat().st_mtime, reverse=True)
        
        for aviso_dir in avisos_dirs:
            numero = aviso_dir.name.replace('aviso_', '')
            mapas = len([f for f in aviso_dir.iterdir() if f.suffix == '.webp'])
            
            avisos.append({
                'numero': numero,
                'fecha': datetime.fromtimestamp(aviso_dir.stat().st_mtime).strftime('%d/%m/%Y %H:%M'),
                'departamentos': ['LIMA', 'LORETO'],  # TODO: extraer de datos reales
                'estado_texto': 'Procesado',
                'estado_color': 'success',
                'num_mapas': mapas,
                'enviado_whatsapp': False  # TODO: implementar tracking WhatsApp
            })
    
    return avisos

def obtener_lista_mapas(aviso_filtro=None):
    """Obtener lista de mapas para galería"""
    mapas = []
    
    if OUTPUT_DIR.exists():
        avisos_dirs = [d for d in OUTPUT_DIR.iterdir() 
                      if d.is_dir() and d.name.startswith('aviso_')]
        
        if aviso_filtro:
            avisos_dirs = [d for d in avisos_dirs if d.name == f'aviso_{aviso_filtro}']
        
        avisos_dirs = sorted(avisos_dirs, key=lambda x: x.stat().st_mtime, reverse=True)
        
        for aviso_dir in avisos_dirs:
            numero_aviso = aviso_dir.name.replace('aviso_', '')
            webp_files = [f for f in aviso_dir.iterdir() if f.suffix == '.webp']
            
            for webp_file in webp_files:
                mapas.append({
                    'id': f"{numero_aviso}_{webp_file.stem}",
                    'nombre': f"{webp_file.stem}.webp",
                    'numero_aviso': numero_aviso,
                    'departamento': webp_file.stem,
                    'url': f"{DOMAIN}/OUTPUT/{aviso_dir.name}/{webp_file.name}",
                    'fecha': datetime.fromtimestamp(aviso_dir.stat().st_mtime).strftime('%d/%m/%Y')
                })
    
    return mapas

def obtener_stats_whatsapp():
    """Estadísticas para página WhatsApp"""
    return {
        'total_contactos': 150,  # TODO: implementar contador real
        'enviados_hoy': 12,
        'pendientes': 3,
        'fallidos': 1
    }

def obtener_avisos_para_whatsapp():
    """Avisos disponibles para envío WhatsApp"""
    avisos = []
    
    if OUTPUT_DIR.exists():
        avisos_dirs = sorted([d for d in OUTPUT_DIR.iterdir() 
                            if d.is_dir() and d.name.startswith('aviso_')], 
                           key=lambda x: x.stat().st_mtime, reverse=True)[:10]
        
        for aviso_dir in avisos_dirs:
            numero = aviso_dir.name.replace('aviso_', '')
            mapas = len([f for f in aviso_dir.iterdir() if f.suffix == '.webp'])
            
            if mapas > 0:  # Solo avisos con mapas
                avisos.append({
                    'numero': numero,
                    'fecha': datetime.fromtimestamp(aviso_dir.stat().st_mtime).strftime('%d/%m/%Y'),
                    'departamentos': ['LIMA', 'LORETO']  # TODO: extraer real
                })
    
    return avisos

def obtener_historial_whatsapp():
    """Historial de envíos WhatsApp"""
    # TODO: implementar con base de datos real
    return []

def obtener_contactos_recientes():
    """Contactos recientes WhatsApp"""
    # TODO: implementar con base de datos real
    return []

# ============================================================================
# RUTAS API - ESTADÍSTICAS
# ============================================================================

@app.route('/api/stats', methods=['GET'])
def api_stats():
    """API endpoint para estadísticas"""
    try:
        stats = obtener_estadisticas()
        return jsonify(stats), 200
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    """Endpoint de salud - verifica que la API está funcionando"""
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    }), 200


@app.route('/OUTPUT/<path:filepath>', methods=['GET'])
def serve_output(filepath):
    """Servir archivos estáticos desde OUTPUT"""
    try:
        return send_from_directory(str(OUTPUT_DIR), filepath, as_attachment=True)
    except Exception as e:
        logger.error(f"Error sirviendo archivo {filepath}: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Archivo no encontrado'}), 404


@app.route('/procesar-aviso', methods=['POST'])
@app.route('/api/procesar-aviso', methods=['POST'])
def procesar_aviso_endpoint():
    """
    Endpoint principal para procesar avisos SENAMHI
    
    Body esperado:
    {
        "numero_aviso": 447,
        "desde_bd": false,
        "json_path": "/path/to/aviso.json"  (opcional)
    }
    
    Respuesta exitosa:
    {
        "status": "success",
        "numero_aviso": 447,
        "output_dir": "/path/to/OUTPUT/aviso_447",
        "mapas": ["departamento1.webp", "departamento2.webp"],
        "archivos_adicionales": ["provincias_afectadas.csv", "distritos_afectados.csv"],
        "timestamp": "2026-01-01T12:00:00"
    }
    """
    try:
        # Validar JSON
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'JSON body requerido'
            }), 400
        
        # Extraer número de aviso
        numero_aviso = data.get('numero_aviso')
        if not numero_aviso:
            return jsonify({
                'status': 'error',
                'message': 'Campo "numero_aviso" requerido'
            }), 400
        
        # Convertir a entero
        try:
            numero_aviso = int(numero_aviso)
        except (ValueError, TypeError):
            return jsonify({
                'status': 'error',
                'message': f'numero_aviso debe ser un entero, recibido: {numero_aviso}'
            }), 400
        
        # Determinar si obtener desde BD o archivo
        desde_bd = data.get('desde_bd', False)
        json_path = data.get('json_path', None)
        
        if json_path:
            json_path = Path(json_path)
            if not json_path.exists():
                return jsonify({
                    'status': 'error',
                    'message': f'Archivo JSON no encontrado: {json_path}'
                }), 400
        
        # Si se solicita desde BD, verificar que exista
        if desde_bd:
            aviso_bd = obtener_aviso_por_numero(numero_aviso)
            if not aviso_bd:
                return jsonify({
                    'status': 'error',
                    'message': f'Aviso {numero_aviso} no encontrado en base de datos'
                }), 404
        
        logger.info(f"Iniciando procesamiento de aviso {numero_aviso} (desde_bd={desde_bd})")
        
        # Verificar si el módulo de procesamiento está disponible
        if not PROCESAR_AVISO_DISPONIBLE:
            return jsonify({
                'status': 'error',
                'message': 'Módulo de procesamiento no disponible. Verifique las dependencias (geopandas, etc.)'
            }), 503
        
        # Procesar aviso
        resultado = procesar_aviso(numero_aviso, desde_bd)
        
        # Obtener lista de mapas generados
        output_dir = OUTPUT_DIR / f"aviso_{numero_aviso}"
        mapas = []
        archivos_adicionales = []
        
        if output_dir.exists():
            for archivo in output_dir.iterdir():
                if archivo.suffix == '.webp':
                    mapas.append(archivo.name)
                elif archivo.suffix == '.csv':
                    archivos_adicionales.append(archivo.name)
        
        logger.info(f"Aviso {numero_aviso} procesado exitosamente. Mapas: {len(mapas)}")
        
        # Generar URLs para los mapas
        mapas_urls = {}
        for mapa in mapas:
            mapa_relative = f"aviso_{numero_aviso}/{mapa}"
            mapas_urls[mapa] = f"{DOMAIN}/OUTPUT/{mapa_relative}"
        
        return jsonify({
            'status': 'success',
            'numero_aviso': numero_aviso,
            'output_dir': str(output_dir),
            'mapas': sorted(mapas),
            'mapas_urls': mapas_urls,
            'archivos_adicionales': sorted(archivos_adicionales),
            'timestamp': datetime.now().isoformat()
        }), 200
    
    except Exception as e:
        logger.error(f"Error procesando aviso: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Error al procesar aviso: {str(e)}'
        }), 500


@app.route('/avisos/<int:numero_aviso>', methods=['GET'])
def obtener_aviso(numero_aviso):
    """
    Endpoint para obtener detalles de un aviso ya procesado
    
    Respuesta:
    {
        "numero_aviso": 447,
        "existe": true,
        "output_dir": "/path/to/OUTPUT/aviso_447",
        "mapas": ["departamento1.webp"],
        "archivos_adicionales": ["provincias_afectadas.csv"]
    }
    """
    try:
        output_dir = OUTPUT_DIR / f"aviso_{numero_aviso}"
        
        if not output_dir.exists():
            return jsonify({
                'status': 'not_found',
                'numero_aviso': numero_aviso,
                'existe': False,
                'message': f'Aviso {numero_aviso} no encontrado'
            }), 404
        
        # Listar archivos
        mapas = []
        archivos_adicionales = []
        
        for archivo in output_dir.iterdir():
            if archivo.suffix == '.webp':
                mapas.append(archivo.name)
            elif archivo.suffix == '.csv':
                archivos_adicionales.append(archivo.name)
        
        return jsonify({
            'status': 'success',
            'numero_aviso': numero_aviso,
            'existe': True,
            'output_dir': str(output_dir),
            'mapas': sorted(mapas),
            'archivos_adicionales': sorted(archivos_adicionales)
        }), 200
    
    except Exception as e:
        logger.error(f"Error obteniendo aviso: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error: {str(e)}'
        }), 500


@app.route('/status', methods=['GET'])
def status():
    """Endpoint para verificar estado general de la API"""
    try:
        # Verificar directorios
        temp_dir = BASE_DIR / "TEMP"
        output_dir = BASE_DIR / "OUTPUT"
        
        return jsonify({
            'status': 'ok',
            'temp_dir': {
                'existe': temp_dir.exists(),
                'ruta': str(temp_dir)
            },
            'output_dir': {
                'existe': output_dir.exists(),
                'ruta': str(output_dir),
                'avisos_procesados': len([d for d in output_dir.iterdir() if d.is_dir()]) if output_dir.exists() else 0
            },
            'timestamp': datetime.now().isoformat()
        }), 200
    
    except Exception as e:
        logger.error(f"Error en status: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error: {str(e)}'
        }), 500


@app.errorhandler(404)
def not_found(error):
    """Manejador de rutas no encontradas"""
    return jsonify({
        'status': 'error',
        'message': 'Ruta no encontrada'
    }), 404


@app.errorhandler(500)
def internal_error(error):
    """Manejador de errores internos"""
    logger.error(f"Error interno: {str(error)}")
    return jsonify({
        'status': 'error',
        'message': 'Error interno del servidor'
    }), 500


@app.route('/api/avisos/nuevos', methods=['GET'])
def api_avisos_nuevos():
    """API para obtener avisos nuevos en las últimas 24 horas"""
    try:
        import psycopg2
        import psycopg2.extras
        from datetime import datetime, timedelta
        
        # Timeout de 5 segundos para conexión
        conn = psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            port=int(os.getenv("DB_PORT", "5432")),
            database=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            connect_timeout=5
        )
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # Buscar avisos con fecha_emision en las últimas 24 horas
        hace_24_horas = datetime.now() - timedelta(hours=24)
        
        query = """
            SELECT COUNT(DISTINCT numero_aviso) as total
            FROM avisos_completos 
            WHERE color IN ('rojo', 'naranja') 
            AND fecha_emision >= %s
        """
        cursor.execute(query, (hace_24_horas,))
        resultado = cursor.fetchone()
        total_nuevos = resultado['total'] if resultado else 0
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'total_nuevos': total_nuevos,
            'fecha_consulta': datetime.now().isoformat()
        })
    except psycopg2.OperationalError as e:
        logger.error(f"Error de conexión BD: {e}")
        return jsonify({
            'status': 'error',
            'message': 'No hay conexión a la base de datos. Intenta más tarde.'
        }), 503
    except Exception as e:
        logger.error(f"Error al obtener avisos nuevos: {e}")
        return jsonify({
            'status': 'error',
            'message': 'Error al consultar avisos nuevos'
        }), 500


@app.route('/api/avisos/<int:numero>/imagenes', methods=['GET'])
def api_avisos_imagenes(numero):
    """API para obtener imágenes y CSV de afectados (para n8n)"""
    try:
        carpeta = BASE_DIR / 'OUTPUT' / f'aviso_{numero}'
        if not carpeta.exists():
            return jsonify({
                'status': 'error',
                'message': f'No hay datos para el aviso {numero}'
            }), 404
        
        # Obtener imágenes WEBP
        imagenes = []
        for archivo in sorted(carpeta.glob('*.webp')):
            imagenes.append({
                'nombre': archivo.name,
                'url': f'/static/output/aviso_{numero}/{archivo.name}',
                'ruta_local': str(archivo)
            })
        
        # Obtener CSVs (departamentos y provincias afectadas)
        archivos_csv = []
        for csv_file in carpeta.glob('*.csv'):
            archivos_csv.append({
                'nombre': csv_file.name,
                'url': f'/static/output/aviso_{numero}/{csv_file.name}',
                'ruta_local': str(csv_file),
                'tipo': 'departamentos' if 'departamentos' in csv_file.name else 'provincias'
            })
        
        return jsonify({
            'status': 'success',
            'numero_aviso': numero,
            'total_imagenes': len(imagenes),
            'total_afectados': len(archivos_csv),
            'imagenes': imagenes,
            'archivos_csv': archivos_csv,
            'mensaje': f'Aviso {numero}: {len(imagenes)} imágenes, {len(archivos_csv)} CSV(s) con departamentos/provincias'
        })
    except Exception as e:
        logger.error(f"Error al obtener imágenes: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


if __name__ == '__main__':
    # Crear directorios si no existen
    (BASE_DIR / "TEMP").mkdir(exist_ok=True)
    (BASE_DIR / "OUTPUT").mkdir(exist_ok=True)
    
    # Configurar exclusiones para el watchdog
    import re
    from werkzeug.serving import run_simple
    
    # Patrones a excluir del watchdog
    exclude_patterns = [
        r'.*TEMP.*',
        r'.*OUTPUT.*',
        r'.*\.pyc',
        r'.*__pycache__.*',
    ]
    
    # Ejecutar con debug en desarrollo
    # Para producción usar gunicorn: gunicorn -w 4 -b 0.0.0.0:5000 app:app
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=False,
        use_reloader=False,
        threaded=True
    )
